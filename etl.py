"""
ETL: Oracle (BO/GDF) -> JSON.GZ -> GitHub Releases
Extrai dados via oracledb, comprime com gzip e salva em data/gz/
para upload no Release 'dados-latest' do GitHub.
Os dashboards consomem os arquivos diretamente do Release.
"""

import os
import json
import gzip
import logging
from datetime import datetime, date, timezone
from decimal import Decimal
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

DB_USER     = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_DSN      = os.getenv("DB_DSN", "10.69.1.118:1521/oraprd06")
CLIENT_PATH = os.getenv("ORACLE_CLIENT_PATH", "").strip()
FCDF_PATH   = os.getenv("FCDF_PATH", "").strip()
DB_MIN  = int(os.getenv("DB_MIN_CONNECTIONS", 1))
DB_MAX  = int(os.getenv("DB_MAX_CONNECTIONS", 5))
DB_INC  = int(os.getenv("DB_INCREMENT_CONNECTIONS", 1))

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "").strip()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / "data"
GZ_DIR      = BASE_DIR / "data" / "gz"
QUERIES_DIR = BASE_DIR / "data" / "queries"
OUTPUT_DIR.mkdir(exist_ok=True)
GZ_DIR.mkdir(exist_ok=True)

SCHEMA_ANO = f"mil{datetime.now().year}"

QUERIES = [
    {
        "file": "receita.json",
        "sql_file": "RECEITA.sql",
    },
    {
        "file": "despesa.json",
        "sql_file": "DESPESA.sql",
    },
    {
        "file": "rcl.json",
        "sql_file": "receita_RCL.sql",
        "transform": "rcl",
    },
    {
        "file": "restos_a_pagar.json",
        "sql_file": "restos_a_pagar.sql",
        "transform": "restos_a_pagar",
    },
    {
        "file": "resultado_primario_nominal.json",
        "sql_file": "resultado_primario_nominal.sql",
        "transform": "resultado_primario_nominal",
    },
    {
        "file": "poupanca_corrente.json",
        "sql_file": "poupanca_corrente.sql",
        "transform": "poupanca_corrente",
    },
]


def serialize(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def fetch(cursor, query):
    cursor.execute(query)
    columns = [col[0].lower() for col in cursor.description]
    return [
        {col: serialize(val) for col, val in zip(columns, row)}
        for row in cursor.fetchall()
    ]


def save_json(filename, data):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "total": len(data),
        "dados": data,
    }
    path = OUTPUT_DIR / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"  {filename} -- {len(data)} registros salvos")


def save_json_gz(filename, data):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "total": len(data),
        "dados": data,
    }
    content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    gz_filename = filename + ".gz"
    path = GZ_DIR / gz_filename
    with gzip.open(path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = path.stat().st_size / 1024
    log.info(f"  {gz_filename} -- {len(data)} registros, {size_kb:.1f} KB comprimido")


def read_sql(filename):
    path = QUERIES_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"Arquivo SQL nao encontrado: {path}")
    sql = path.read_text(encoding="utf-8")
    lines = [line for line in sql.splitlines() if not line.strip().startswith("--")]
    return "\n".join(lines).replace("{SCHEMA_ANO}", SCHEMA_ANO).strip()


def resolve_query(item):
    return read_sql(item["sql_file"])


# RCL Aggregation
# Regras: tools/regra_rcl.txt (ContDF/SEEC)
# FCDF: data/UFIS-FCDFDespesadePessoal.xlsx (UFIS/SIAFE)

FCDF_CLASS6_CONTAS = {622130300, 622130400, 622130500, 622130600, 622130700}

# RREO Anexo 06 — Juros Ativos (XXXVI) e Passivos (XXXVII)
_XXXVI_CC5 = {
    "44121","44123","44124","44125","44131","44133","44134","44135","44141",
    "44211","44213","44214","44215","44221","44261","44263","44264","44265",
    "44511","44521","44611","44613","44614","44615","44621","44623","44624","44625",
}
_XXXVI_CC7 = {
    "4411199","4431101","4431199","4431301","4431401","4431501",
    "4432101","4433101","4433199","4433301","4433401","4433501",
    "4434101","4435101","4435301","4435401","4435501",
}
_XXXVI_FULL = {"443910170", "443930170", "443930171"}

_XXXVII_CC5 = {
    "34111","34113","34114","34115","34121","34131","34133","34134","34135","34141",
    "34181","34183","34184","34185","34191","34211","34213","34214","34215","34221",
    "34261","34263","34264","34265","34511","34521","34611","34613","34614","34615",
    "34911","34913","34914","34915",
}
_XXXVII_CC7 = {
    "3425202","3431101","3431301","3431401","3431501","3432101","3433101",
    "3433301","3433401","3433501","3434101","3435101","3435301","3435401","3435501",
}
_XXXVII_FULL = {"343910170", "343930170", "343930171"}


def _rcl_class_orc(c, cofonte, cofontefederal):
    if 11125000 <= c <= 11125099: return "iptu"
    if 11130000 <= c <= 11139999: return "ir"
    if 11125100 <= c <= 11125199: return "ipva"
    if 11125200 <= c <= 11125299: return "itcd"
    if 11125300 <= c <= 11125399: return "itbi"
    if (11145010 <= c <= 11145099) or (11145200 <= c <= 11145299): return "icms"
    if 11145100 <= c <= 11145199: return "iss"
    if 11190000 <= c <= 11199999: return "outros_impostos"
    if 11200000 <= c <= 11299999: return "taxas"
    if 12000000 <= c <= 12999999: return "contribuicoes"
    if 13200000 <= c <= 13299999: return "rend_aplic"
    if (13100000 <= c <= 13199999) or (13300000 <= c <= 13999999): return "outras_patrimoniais"
    if 14000000 <= c <= 14999999: return "agropecuaria"
    if 15000000 <= c <= 15999999: return "industrial"
    if 16000000 <= c <= 16999999: return "servicos"
    if 17115000 <= c <= 17115099: return "fpe"
    if 17115100 <= c <= 17115199: return "fpm"
    if 17115200 <= c <= 17115299: return "itr_trans"
    if 17115300 <= c <= 17115399: return "lc61"
    if (17515000 <= c <= 17515099) or (17155200 <= c <= 17155299): return "fundeb_trans"
    if (17115400 <= c <= 17155199) or (17155300 <= c <= 17514999) or (17515100 <= c <= 17999999):
        return "outras_transf"
    if 19000000 <= c <= 19999999: return "outras_correntes"
    return None


def _rcl_deducao(c):
    if 12150000 <= c <= 12159999: return "contrib_servidor"
    if 19990300 <= c <= 19990399: return "comp_financeira"
    if 13210400 <= c <= 13210499: return "rend_prev"
    if 17515000 <= c <= 17515099: return "ded_fundeb"
    return None


def _is_xxxvi(cc):
    return cc[:5] in _XXXVI_CC5 or cc[:7] in _XXXVI_CC7 or cc in _XXXVI_FULL


def _is_xxxvii(cc):
    return cc[:5] in _XXXVII_CC5 or cc[:7] in _XXXVII_CC7 or cc in _XXXVII_FULL


def _rcl_emenda(cofonte, cofontefederal):
    if (732000000 <= cofonte <= 732999999 or
            738000000 <= cofonte <= 738999999 or
            706000000 <= cofonte <= 706999999):
        return "emendas_ind"
    if (733000000 <= cofonte <= 733999999 or
            739000000 <= cofonte <= 739999999):
        return "emendas_bancada"
    if cofontefederal == 1604:
        return "agentes_com"
    return None


def load_fcdf_data(base_dir):
    if FCDF_PATH:
        path = Path(FCDF_PATH)
    else:
        path = base_dir / "data" / "UFIS-FCDFDespesadePessoal.xlsx"
    if not path.exists():
        log.warning(f"Planilha FCDF nao encontrada: {path}")
        return {"realizados": {}, "previsao": {}}
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl nao instalado. Execute: pip install openpyxl")
        return {"realizados": {}, "previsao": {}}
    from collections import defaultdict
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    realizados = defaultdict(lambda: {"total": 0.0, "pessoal": 0.0})
    previsao   = defaultdict(lambda: {"total": 0.0, "pessoal": 0.0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        try:
            conta  = int(row[2])
            grupo  = int(row[4] or 0)
            elem   = str(row[5]).strip() if row[5] is not None else ""
            subit  = int(row[6]) if row[6] is not None else 0
            mes    = int(row[7] or 0)
            ano    = int(row[8] or 0)
            vadeb  = float(row[9] or 0)
            vacred = float(row[10] or 0)
        except (ValueError, TypeError):
            continue
        if not mes or not ano:
            continue
        is_pessoal = (grupo == 1) or (grupo == 3 and elem == "85" and subit == 1)
        sc = str(conta)
        if sc.startswith("6") and conta in FCDF_CLASS6_CONTAS:
            saldo = vacred - vadeb
            realizados[(mes, ano)]["total"]   += saldo
            if is_pessoal:
                realizados[(mes, ano)]["pessoal"] += saldo
        elif sc.startswith("5"):
            saldo = vadeb - vacred
            previsao[ano]["total"]   += saldo
            if is_pessoal:
                previsao[ano]["pessoal"] += saldo
    wb.close()
    log.info(f"  FCDF: {len(realizados)} meses realizados, {len(previsao)} anos previsao")
    return {"realizados": dict(realizados), "previsao": dict(previsao)}


def build_rcl_data(rows):
    from collections import defaultdict
    MESES_PT = ["jan","fev","mar","abr","mai","jun",
                "jul","ago","set","out","nov","dez"]
    ano_atual = datetime.now().year
    PREV_EXCL = {521120101, 521220101, 521220201}
    prev_receita  = defaultdict(float)
    prev_deducoes = defaultdict(float)
    prev_emendas  = defaultdict(float)
    real_receita  = defaultdict(lambda: defaultdict(float))
    real_deducoes = defaultdict(lambda: defaultdict(float))
    real_emendas  = defaultdict(lambda: defaultdict(float))
    meses_oracle  = set()

    for r in rows:
        saldo = float(r.get("saldo") or 0)
        if not saldo:
            continue
        cc_raw = r.get("cocontacontabil")
        cc = int(str(cc_raw).strip()) if cc_raw is not None else 0
        class_orc_raw = r.get("class_orc") or ""
        try:
            c_int = int(str(class_orc_raw).strip())
        except ValueError:
            continue
        if not str(class_orc_raw).strip().startswith("1"):
            continue
        mes = int(r.get("inmes") or 0)
        ano = int(r.get("coexercicio") or 0)
        if not mes or not ano:
            continue
        cofonte_raw   = r.get("cofonte")
        cofederal_raw = r.get("cofontefederal")
        try:
            cofonte   = int(str(cofonte_raw).strip()) if cofonte_raw else 0
            cofederal = int(str(cofederal_raw).strip()) if cofederal_raw else 0
        except (ValueError, TypeError):
            cofonte = cofederal = 0

        is_prev = (521100000 <= cc <= 521299999)
        is_real = (621200000 <= cc <= 621399999 and cc != 621310100)

        if is_prev:
            is_excl = (cc in PREV_EXCL)
            if not is_excl:
                key = _rcl_class_orc(c_int, cofonte, cofederal)
                if key and ano == ano_atual:
                    prev_receita[key] += saldo
                ded = _rcl_deducao(c_int)
                # ded_fundeb na previsao vem APENAS das contas excluidas (PREV_EXCL)
                if ded and ded != "ded_fundeb" and ano == ano_atual:
                    prev_deducoes[ded] += saldo
                em = _rcl_emenda(cofonte, cofederal)
                if em and ano == ano_atual:
                    prev_emendas[em] += saldo
            else:
                if 17515000 <= c_int <= 17515099 and ano == ano_atual:
                    prev_deducoes["ded_fundeb"] += saldo
        elif is_real:
            meses_oracle.add((mes, ano))
            key = _rcl_class_orc(c_int, cofonte, cofederal)
            if key:
                real_receita[(mes, ano)][key] += saldo
            ded = _rcl_deducao(c_int)
            if ded:
                real_deducoes[(mes, ano)][ded] += saldo
            em = _rcl_emenda(cofonte, cofederal)
            if em:
                real_emendas[(mes, ano)][em] += saldo

    if not meses_oracle:
        log.warning("RCL: nenhum dado realizado encontrado no SQL.")
        return {}

    # ref_mes/ref_ano: ultimo mes fechado conforme {SCHEMA_ANO}.mesfechado.
    # max_mes_fechado vem como coluna escalar no SQL (mesmo valor em todas as linhas).
    # Fallback: se NULL (mesfechado vazio no ano corrente), usa max dos realizados Oracle.
    mmf_raw = next((r.get("max_mes_fechado") for r in rows if r.get("max_mes_fechado") is not None), None)
    if mmf_raw is not None and 1 <= int(mmf_raw) <= 12:
        ref_mes = int(mmf_raw)
        ref_ano = ano_atual
        log.info(f"  RCL: ultimo mes fechado (mesfechado) = {ref_mes:02d}/{ref_ano}")
    else:
        # Fallback: max dos realizados Oracle
        max_mes, max_ano = max(meses_oracle)
        ref_mes, ref_ano = max_mes, max_ano
        log.warning(f"  RCL: mesfechado vazio/invalido, fallback Oracle max = {ref_mes:02d}/{ref_ano}")

    ultimos12 = []
    m, a = ref_mes, ref_ano
    for _ in range(12):
        ultimos12.insert(0, (m, a))
        m -= 1
        if m == 0:
            m = 12
            a -= 1

    # Janela completa: Jan/ano_anterior até ref_mes/ref_ano (suporte a bimestres)
    ano_anterior = ref_ano - 1
    todas_colunas_ma = ([(mm, ano_anterior) for mm in range(1, 13)] +
                        [(mm, ref_ano) for mm in range(1, ref_mes + 1)])

    janela_padrao = [f"{m},{a}" for m, a in ultimos12]
    colunas  = [f"{m},{a}" for m, a in todas_colunas_ma]
    rotulos  = [f"{MESES_PT[m-1]}/{str(a)[2:]}" for m, a in todas_colunas_ma]
    log.info(f"  RCL: colunas {rotulos[0]} -> {rotulos[-1]} ({len(colunas)} meses)")

    def monta_linha(src, key):
        linha = {}
        for m, a in todas_colunas_ma:
            col = f"{m},{a}"
            linha[col] = src.get((m, a), {}).get(key, 0.0)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    def soma_linhas(linhas_dict, keys):
        linha = {}
        for col in colunas:
            linha[col] = sum(linhas_dict.get(k, {}).get(col, 0.0) for k in keys)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    KEYS_ATOMICAS = [
        "iptu","ir","ipva","itcd","itbi","icms","iss","outros_impostos",
        "taxas","itr","contribuicoes","rend_aplic","outras_patrimoniais",
        "agropecuaria","industrial","servicos",
        "fpe","fpm","itr_trans","lc61","fundeb_trans","outras_transf",
        "outras_correntes",
    ]
    IMPOSTOS_KEYS = {"iptu","ir","ipva","itcd","itbi","icms","iss","outros_impostos","taxas","itr"}
    PATRIM_KEYS   = {"rend_aplic","outras_patrimoniais"}
    TRANSF_KEYS   = {"fpe","fpm","itr_trans","lc61","fundeb_trans","outras_transf"}
    CORR_KEYS     = IMPOSTOS_KEYS | PATRIM_KEYS | TRANSF_KEYS | {
                        "contribuicoes","agropecuaria","industrial","servicos","outras_correntes"}

    linhas = {}
    for key in KEYS_ATOMICAS:
        linhas[key] = monta_linha(real_receita, key)
    linhas["itr"] = {col: 0.0 for col in colunas}
    linhas["itr"]["_total"] = 0.0

    linhas["impostos"]           = soma_linhas(linhas, IMPOSTOS_KEYS)
    linhas["patrimonial"]        = soma_linhas(linhas, PATRIM_KEYS)
    linhas["transferencias"]     = soma_linhas(linhas, TRANSF_KEYS)
    linhas["receitas_correntes"] = soma_linhas(linhas, CORR_KEYS)

    for dk in ("contrib_servidor","comp_financeira","rend_prev","ded_fundeb"):
        linhas[dk] = monta_linha(real_deducoes, dk)
    linhas["deducoes"] = soma_linhas(linhas,
        {"contrib_servidor","comp_financeira","rend_prev","ded_fundeb"})

    fcdf      = load_fcdf_data(BASE_DIR)
    real_fcdf = fcdf.get("realizados", {})
    prev_fcdf = fcdf.get("previsao", {})

    def monta_fcdf(campo):
        linha = {}
        for m, a in todas_colunas_ma:
            col = f"{m},{a}"
            linha[col] = real_fcdf.get((m, a), {}).get(campo, 0.0)
        linha["_total"] = sum(linha.get(c, 0.0) for c in janela_padrao)
        return linha

    linhas["fcdf_total"]   = monta_fcdf("total")
    linhas["fcdf_pessoal"] = monta_fcdf("pessoal")
    linhas["fcdf"] = {}
    for col in colunas:
        linhas["fcdf"][col] = (linhas["fcdf_total"].get(col, 0)
                               - linhas["fcdf_pessoal"].get(col, 0))
    linhas["fcdf"]["_total"] = (linhas["fcdf_total"]["_total"]
                                - linhas["fcdf_pessoal"]["_total"])

    linhas["rcl"] = {}
    for col in colunas:
        linhas["rcl"][col] = (linhas["receitas_correntes"].get(col, 0)
                              - linhas["deducoes"].get(col, 0)
                              + linhas["fcdf"].get(col, 0))
    linhas["rcl"]["_total"] = (linhas["receitas_correntes"]["_total"]
                               - linhas["deducoes"]["_total"]
                               + linhas["fcdf"]["_total"])

    for em in ("emendas_ind","emendas_bancada","agentes_com"):
        linhas[em] = monta_linha(real_emendas, em)
    linhas["outras_ded"] = {col: 0.0 for col in colunas}
    linhas["outras_ded"]["_total"] = 0.0

    linhas["rcl_endiv"] = {}
    for col in colunas:
        linhas["rcl_endiv"][col] = (linhas["rcl"].get(col, 0)
                                    - linhas["emendas_ind"].get(col, 0))
    linhas["rcl_endiv"]["_total"] = (linhas["rcl"]["_total"]
                                     - linhas["emendas_ind"]["_total"])

    linhas["rcl_pessoal"] = {}
    for col in colunas:
        linhas["rcl_pessoal"][col] = (
            linhas["rcl_endiv"].get(col, 0)
            - linhas["emendas_bancada"].get(col, 0)
            - linhas["agentes_com"].get(col, 0)
            - linhas["outras_ded"].get(col, 0))
    linhas["rcl_pessoal"]["_total"] = (
        linhas["rcl_endiv"]["_total"]
        - linhas["emendas_bancada"]["_total"]
        - linhas["agentes_com"]["_total"]
        - linhas["outras_ded"]["_total"])

    def pv(key): return prev_receita.get(key, 0.0)
    def pv_g(keys): return sum(prev_receita.get(k, 0.0) for k in keys)

    previsao = {}
    for k in KEYS_ATOMICAS:
        previsao[k] = pv(k)
    previsao["itr"]                = 0.0
    previsao["impostos"]           = pv_g(IMPOSTOS_KEYS)
    previsao["patrimonial"]        = pv_g(PATRIM_KEYS)
    previsao["transferencias"]     = pv_g(TRANSF_KEYS)
    previsao["receitas_correntes"] = pv_g(CORR_KEYS)

    for dk in ("contrib_servidor","comp_financeira","rend_prev","ded_fundeb"):
        previsao[dk] = prev_deducoes.get(dk, 0.0)
    previsao["deducoes"] = sum(prev_deducoes.values())

    pf = prev_fcdf.get(ano_atual, {})
    previsao["fcdf_total"]   = pf.get("total", 0.0)
    previsao["fcdf_pessoal"] = pf.get("pessoal", 0.0)
    previsao["fcdf"]         = previsao["fcdf_total"] - previsao["fcdf_pessoal"]

    previsao["rcl"] = (previsao["receitas_correntes"]
                       - previsao["deducoes"]
                       + previsao["fcdf"])

    for em in ("emendas_ind","emendas_bancada","agentes_com"):
        previsao[em] = prev_emendas.get(em, 0.0)
    previsao["outras_ded"] = 0.0

    previsao["rcl_endiv"]   = previsao["rcl"] - previsao["emendas_ind"]
    previsao["rcl_pessoal"] = (previsao["rcl_endiv"]
                               - previsao["emendas_bancada"]
                               - previsao["agentes_com"]
                               - previsao["outras_ded"])

    return {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "ano": ano_atual,
        "ref_mes": ref_mes,
        "ref_ano": ref_ano,
        "colunas": colunas,
        "rotulos": rotulos,
        "janela_padrao": janela_padrao,
        "linhas": linhas,
        "previsao": previsao,
    }


def build_restos_a_pagar_data(rows):
    """
    Agrega as linhas brutas do SQL por (ano, coug, cocontacontabil, cat, gnd, inmes)
    e retorna lista de registros prontos para o dashboard.
    Saldo = VACREDITO - VADEBITO (contas classe 6 de execução de RAP).
    """
    agg  = {}   # chave -> saldo acumulado
    meta = {}   # chave -> campos descritivos

    for r in rows:
        try:
            cc    = int(str(r.get("cocontacontabil") or 0).strip())
            saldo = float(r.get("vacredito") or 0) - float(r.get("vadebito") or 0)
            inmes = int(r.get("inmes") or 0)
            ano   = int(r.get("ano") or 0)
        except (ValueError, TypeError):
            continue
        if not ano:
            continue

        coug = str(r.get("coug") or "").strip()
        cat  = str(r.get("cocategoriaeconomica") or "").strip()
        gnd  = str(r.get("cognd") or "").strip()
        key  = (ano, coug, cc, cat, gnd, inmes)

        agg[key]  = agg.get(key, 0.0) + saldo
        meta[key] = {
            "noug":  str(r.get("noug")  or "").strip(),
            "nocat": str(r.get("nocategoriaeconomica") or "").strip(),
            "nognd": str(r.get("nognd") or "").strip(),
        }

    registros = []
    for (ano, coug, cc, cat, gnd, inmes), saldo in agg.items():
        m = meta[(ano, coug, cc, cat, gnd, inmes)]
        registros.append({
            "ano":             ano,
            "coug":            coug,
            "noug":            m["noug"],
            "cocontacontabil": cc,
            "cat":             cat,
            "nocat":           m["nocat"],
            "gnd":             gnd,
            "nognd":           m["nognd"],
            "saldo":           round(saldo, 2),
            "inmes":           inmes,
        })

    log.info(f"  Restos a Pagar: {len(registros)} registros agregados")
    return registros


def _supabase_upsert(table, payload, on_conflict, batch_size=1000):
    """Função genérica de upsert no Supabase via REST API."""
    import urllib.request, urllib.error
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates",
    }
    total = 0
    for i in range(0, len(payload), batch_size):
        body = json.dumps(payload[i:i+batch_size], ensure_ascii=False).encode("utf-8")
        req  = urllib.request.Request(url, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                resp.read()
            total += len(payload[i:i+batch_size])
        except urllib.error.HTTPError as e:
            body_err = e.read().decode("utf-8", errors="replace")
            log.error(f"  Supabase [{table}] lote {i}: HTTP {e.code} - {body_err}")
            raise
    return total


def upsert_restos_a_pagar_supabase(registros):
    """Envia restos a pagar para o Supabase."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert restos_a_pagar.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        payload = [{**r, "atualizado_em": atualizado_em} for r in registros]
        total = _supabase_upsert("restos_a_pagar", payload,
                                  "ano,coug,cocontacontabil,cat,gnd,inmes")
        log.info(f"  Supabase: {total} registros enviados para restos_a_pagar.")
    except Exception as e:
        log.error(f"  Supabase restos_a_pagar falhou: {type(e).__name__}: {e}")


def upsert_receita_supabase(data):
    """Envia receita orçamentária para o Supabase."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert receita.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        payload = []
        for r in data:
            row = {k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
            row["atualizado_em"] = atualizado_em
            payload.append(row)
        total = _supabase_upsert("receita", payload,
                                  "coexercicio,inmes,coug,cocontacontabil,cocontacorrente")
        log.info(f"  Supabase: {total} registros enviados para receita.")
    except Exception as e:
        log.error(f"  Supabase receita falhou: {type(e).__name__}: {e}")


def upsert_despesa_supabase(data):
    """Envia despesa orçamentária para o Supabase (com deduplicação por chave única)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert despesa.")
        return
    try:
        atualizado_em = datetime.now(timezone.utc).isoformat()
        # Deduplicar pela chave única, somando valores financeiros
        agg = {}
        meta = {}
        for r in data:
            row = {k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
            row["cofonte"] = row.get("cofonte") or ""
            key = (row.get("coexercicio"), row.get("inmes"), row.get("coug"),
                   row.get("cocontacontabil"), row.get("despesa"), row.get("cofonte"))
            if key not in agg:
                agg[key]  = {"vadebito": 0.0, "vacredito": 0.0, "saldo": 0.0}
                meta[key] = row
            agg[key]["vadebito"]  += float(row.get("vadebito")  or 0)
            agg[key]["vacredito"] += float(row.get("vacredito") or 0)
            agg[key]["saldo"]     += float(row.get("saldo")     or 0)

        payload = []
        for key, vals in agg.items():
            row = {**meta[key], **vals, "atualizado_em": atualizado_em}
            payload.append(row)

        log.info(f"  Despesa: {len(data)} linhas Oracle -> {len(payload)} registros únicos.")
        total = _supabase_upsert("despesa", payload,
                                  "coexercicio,inmes,coug,cocontacontabil,despesa,cofonte",
                                  batch_size=1000)
        log.info(f"  Supabase: {total} registros enviados para despesa.")
    except Exception as e:
        log.error(f"  Supabase despesa falhou: {type(e).__name__}: {e}")


def upsert_resultado_primario_nominal_supabase(D_obj):
    """Envia resultado_primario_nominal para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert resultado_primario_nominal.")
        return
    try:
        ano = datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("resultado_primario_nominal", payload, "ano")
        log.info(f"  Supabase: resultado_primario_nominal {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase resultado_primario_nominal falhou: {type(e).__name__}: {e}")


def upsert_rcl_supabase(D_obj):
    """Envia RCL para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert rcl.")
        return
    try:
        ano = D_obj.get("ano") or datetime.now().year
        payload = [{
            "ano":          ano,
            "dados":        D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("rcl", payload, "ano")
        log.info(f"  Supabase: RCL {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase rcl falhou: {type(e).__name__}: {e}")


def save_restos_a_pagar_gz(registros):
    payload = {
        "atualizado_em": datetime.now(timezone.utc).isoformat(),
        "registros": registros,
    }
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "restos_a_pagar.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  restos_a_pagar.json.gz -- {len(registros)} registros, {size_kb:.1f} KB")


def save_resultado_primario_nominal_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "resultado_primario_nominal.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  resultado_primario_nominal.json.gz -- {size_kb:.1f} KB")


def save_rcl_gz(D_obj):
    content = json.dumps(D_obj, ensure_ascii=False, separators=(",",":")).encode("utf-8")
    gz_path = GZ_DIR / "rcl.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  rcl.json.gz -- {len(D_obj.get('colunas', []))} meses, {size_kb:.1f} KB")


def build_poupanca_corrente_data(rows):
    """
    Computa Poupança Corrente (Art. 167-A, CF) com janela móvel de 12 meses.
    Formula: (Desp. Liquidadas 12m + RPNP Inscrito - RPNP Cancelado) / Rec. Corrente 12m
    """
    ano_atual    = datetime.now().year
    ano_anterior = ano_atual - 1

    rec      = {ano_anterior: {m: 0.0 for m in range(1, 13)},
                ano_atual:    {m: 0.0 for m in range(1, 13)}}
    desp_liq = {ano_anterior: {m: 0.0 for m in range(1, 13)},
                ano_atual:    {m: 0.0 for m in range(1, 13)}}
    rpnp_ins = {ano_anterior: 0.0, ano_atual: 0.0}
    rpnp_can = {ano_anterior: {m: 0.0 for m in range(1, 13)},
                ano_atual:    {m: 0.0 for m in range(1, 13)}}
    max_mes  = 0

    for r in rows:
        cc   = str(r.get("cocontacontabil") or "").strip()
        ccor = str(r.get("cocontacorrente") or "").strip()
        nat  = str(r.get("conatureza")      or "").strip()
        mes  = int(r.get("inmes") or 0)
        ano  = int(r.get("coexercicio") or 0)
        vacr = float(r.get("vacredito") or 0)
        vad  = float(r.get("vadebito")  or 0)
        mmf  = r.get("max_mes_fechado")
        if mmf is not None:
            try:
                max_mes = max(max_mes, int(mmf))
            except (ValueError, TypeError):
                pass
        if ano not in rec:
            continue
        try:
            cc_int = int(cc)
        except (ValueError, TypeError):
            continue
        nat2 = nat[1:2] if len(nat) >= 2 else ""
        val  = vacr - vad

        if (621200000 <= cc_int <= 621390199
                and ccor[:1] in ('1', '7')
                and 1 <= mes <= 12):
            rec[ano][mes] += val
        elif (cc[:7] in ('6221303', '6221304', '6221307')
                and nat2 in ('1', '2', '3')
                and 1 <= mes <= 12):
            desp_liq[ano][mes] += val
        elif (cc_int in (631100000, 631200000)
                and nat2 in ('1', '2', '3')
                and mes == 0):
            rpnp_ins[ano] += val
        elif (cc_int == 631900000
                and nat2 in ('1', '2', '3', '7')
                and 1 <= mes <= 12):
            rpnp_can[ano][mes] += val

    if max_mes == 0:
        max_mes = next((m for m in range(12, 0, -1)
                        if rec[ano_atual][m] != 0.0), 1)

    por_mes = {}
    for mes in range(1, max_mes + 1):
        rec_12m  = (sum(rec[ano_anterior][m]      for m in range(mes + 1, 13))
                  + sum(rec[ano_atual][m]          for m in range(1, mes + 1)))
        desp_12m = (sum(desp_liq[ano_anterior][m] for m in range(mes + 1, 13))
                  + sum(desp_liq[ano_atual][m]     for m in range(1, mes + 1)))
        rpnp_i   = rpnp_ins[ano_anterior]
        rpnp_c   = sum(rpnp_can[ano_atual][m] for m in range(1, mes + 1))
        desp_cor = desp_12m + rpnp_i - rpnp_c
        pct      = round(desp_cor / rec_12m * 100, 2) if rec_12m else None
        por_mes[str(mes)] = {
            "rec_corrente_12m":    round(rec_12m,  2),
            "desp_liquidadas_12m": round(desp_12m, 2),
            "rpnp_inscrito":       round(rpnp_i,   2),
            "rpnp_cancelado":      round(rpnp_c,   2),
            "desp_corrente_12m":   round(desp_cor, 2),
            "poupanca_pct":        pct,
        }

    log.info(f"  Poupança Corrente: max_mes={max_mes}, meses={list(por_mes.keys())}")
    return {
        "ano_atual":  ano_atual,
        "max_mes":    max_mes,
        "limite_pct": 95.0,
        "por_mes":    por_mes,
    }


def save_poupanca_corrente_gz(D_obj):
    payload = {"atualizado_em": datetime.now(timezone.utc).isoformat()}
    payload.update(D_obj)
    content = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    gz_path = GZ_DIR / "poupanca_corrente.json.gz"
    with gzip.open(gz_path, "wb", compresslevel=9) as f:
        f.write(content)
    size_kb = gz_path.stat().st_size / 1024
    log.info(f"  poupanca_corrente.json.gz -- {size_kb:.1f} KB")


def upsert_poupanca_corrente_supabase(D_obj):
    """Envia poupanca_corrente para o Supabase como JSONB (1 linha por ano)."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning("  Supabase: nao configurado. Pulando upsert poupanca_corrente.")
        return
    try:
        ano = D_obj.get("ano_atual") or datetime.now().year
        payload = [{
            "ano":           ano,
            "dados":         D_obj,
            "atualizado_em": datetime.now(timezone.utc).isoformat(),
        }]
        total = _supabase_upsert("poupanca_corrente", payload, "ano")
        log.info(f"  Supabase: poupanca_corrente {ano} enviada ({total} linha).")
    except Exception as e:
        log.error(f"  Supabase poupanca_corrente falhou: {type(e).__name__}: {e}")


def build_resultado_primario_nominal_data(rows):
    """
    Computa o RREO Anexo 06 (Acima da Linha) com todos os sub-níveis,
    por mês individual. O browser soma os meses selecionados.
    Regras: tools/regra_ressultado_primario_nominal.txt
    """
    def _fmt3(arr):
        return {"a": round(arr[0], 2), "b": round(arr[1], 2), "c": round(arr[2], 2),
                "total": round(arr[0] + arr[1] + arr[2], 2)}

    REC_KEYS = [
        # Principais (já existiam)
        "I","II","III","V","VII","VIII","IX","X","XI","XII","XIV",
        # Sub-níveis correntes EXCETO RPPS
        "icms","ipva","itcd","iptu","itbi","iss","ir","outros_impostos","taxas",
        "contribuicoes",
        "outras_patrimoniais",
        "fpe","fpm","itr_trans","lc87","lc61","fundeb","outras_transf_corr",
        "correntes_restantes",
        # Sub-níveis capital EXCETO RPPS
        "outras_alienacoes","convenios","outras_transf_cap","outras_cap_prim",
    ]
    DEP_KEYS = [
        # Principais (já existiam)
        "XVIII","XIX","XXI","XXIII","XXIV","XXV","XXVI","XXVII","XXIX","XXX",
        # Sub-níveis
        "pessoal","outras_correntes_dep","investimentos","demais_inv",
    ]

    M_rec = {m: {k: 0.0 for k in REC_KEYS} for m in range(1, 13)}
    M_dep = {m: {k: [0.0, 0.0, 0.0] for k in DEP_KEYS} for m in range(1, 13)}
    M_jur = {m: {"XXXVI": 0.0, "XXXVII": 0.0} for m in range(1, 13)}
    P_rec = {k: 0.0 for k in REC_KEYS}

    _APLIC_FIN = {"132101","132102","132103","132104","132105","132999",
                  "732101","732102","732103","732104","732105","732999"}
    _FUNDEB_4  = {"1715","1751","7715","7751"}

    def _classifica_rec(val, co, cc_corr, is_exceto_rpps, bucket):
        co2, co3, co4, co6, co7 = co[:2], co[:3], co[:4], co[:6], co[:7]
        try:
            co_int = int(co)
        except (ValueError, TypeError):
            co_int = 0
        is_corrente = co2 in {"11","12","13","14","15","16","17","19",
                               "71","72","73","74","75","76","77","79"}
        is_capital  = co2 in {"21","22","23","24","29","81","82","83","84","89"}
        is_aplic    = cc_corr[:6] in _APLIC_FIN
        is_out_fin  = (co4 in {"1944","7944"} or
                       co6 in {"164101","164103","199911","764101","764103","799911"} or
                       co7 in {"1922012","1922064","1922142","1999993",
                                "7922012","7922064","7922142","7999993"})

        if is_exceto_rpps:
            if is_corrente:
                bucket["I"] += val
                if is_aplic:
                    bucket["II"] += val
                elif is_out_fin:
                    bucket["III"] += val
                # Sub-categorias
                if co2 in {"11","71"}:
                    if   (11145010 <= co_int <= 11145099) or (11145200 <= co_int <= 11145299):
                        bucket["icms"]          += val
                    elif co6 in {"111251","711251"}:
                        bucket["ipva"]          += val
                    elif 11125200 <= co_int <= 11125299:
                        bucket["itcd"]          += val
                    elif 11125000 <= co_int <= 11125099:
                        bucket["iptu"]          += val
                    elif 11125300 <= co_int <= 11125399:
                        bucket["itbi"]          += val
                    elif co7 in {"1114511","1114512","7114511","7114512"}:
                        bucket["iss"]           += val
                    elif co6 in {"111303","711303"}:
                        bucket["ir"]            += val
                    elif co4 in {"1119"}:
                        bucket["outros_impostos"] += val
                    elif co3 in {"112","712"}:
                        bucket["taxas"]         += val
                elif co2 in {"12","72"}:
                    bucket["contribuicoes"]     += val
                elif co2 in {"13","73"}:
                    if not is_aplic:
                        bucket["outras_patrimoniais"] += val
                elif co2 in {"17","77"}:
                    if   17115000 <= co_int <= 17115099:
                        bucket["fpe"]            += val
                    elif 17115100 <= co_int <= 17115199:
                        bucket["fpm"]            += val
                    elif 17115200 <= co_int <= 17115299:
                        bucket["itr_trans"]      += val
                    elif 17115300 <= co_int <= 17115399:
                        bucket["lc61"]           += val
                    elif co4 in _FUNDEB_4:
                        bucket["fundeb"]         += val
                    else:
                        bucket["outras_transf_corr"] += val
                elif co2 in {"14","15","16","19","74","75","76","79"}:
                    if not is_out_fin:
                        bucket["correntes_restantes"] += val
            elif is_capital:
                bucket["VII"] += val
                if   co2 in {"21","81"}:
                    bucket["VIII"] += val
                elif co2 in {"23","83"}:
                    bucket["IX"]   += val
                elif co6 in {"221101","821101"}:
                    bucket["X"]    += val
                elif co6 in {"221102","821102"}:
                    bucket["XI"]   += val
                elif co2 in {"22","82"}:
                    bucket["outras_alienacoes"] += val
                elif co2 in {"24","84"}:
                    if (co4 in {"2414","2422","2432","8414","8422","8432"} or
                            co6 in {"244150","244151","844150","844151"}):
                        bucket["convenios"]          += val
                    else:
                        bucket["outras_transf_cap"]  += val
                elif co3 in {"292","293","294","892","893","894"}:
                    bucket["XII"] += val
                elif co3 in {"291","299","891","899"}:
                    bucket["outras_cap_prim"] += val
        else:
            if is_corrente and not (is_aplic or is_out_fin):
                bucket["V"] += val
            elif ((co2 in {"22","24","82","84"} or co3 in {"291","299","891","899"}) and
                  co6 not in {"221101","221102","821101","821102"}):
                bucket["XIV"] += val

    for r in rows:
        cc = str(r.get("cocontacontabil") or "").strip()
        if not cc:
            continue
        mes     = int(r.get("inmes") or 0)
        vacred  = float(r.get("vacredito") or 0)
        vadeb   = float(r.get("vadebito")  or 0)
        cf      = str(r.get("cofontefederal") or "").strip()
        co      = str(r.get("coclasseorc")    or "").strip()
        cc_corr = str(r.get("cocontacorrente") or "").strip()
        na      = str(r.get("conatureza")     or "").strip()
        func    = str(r.get("cofuncao")       or "").strip()
        is_rpps        = cf[1:4] in {"800","801","802"} if len(cf) >= 4 else False
        is_exceto_rpps = not is_rpps
        cc4 = cc[:4]
        cc7 = cc[:7]

        if cc4 in {"6212","6213"}:
            if not mes or mes > 12:
                continue
            _classifica_rec(vacred - vadeb, co, cc_corr, is_exceto_rpps, M_rec[mes])

        elif cc4 in {"5211","5212"}:
            _classifica_rec(vadeb - vacred, co, cc_corr, is_exceto_rpps, P_rec)

        else:
            if not mes or mes > 12:
                continue
            if   cc7 == "6221304":                col, val = 0, vacred - vadeb
            elif cc4 == "6322":                   col, val = 1, vacred - vadeb
            elif cc in {"631400000","631820000"}:  col, val = 2, vacred - vadeb
            elif _is_xxxvi(cc)  and is_exceto_rpps:
                M_jur[mes]["XXXVI"]  += vacred - vadeb; continue
            elif _is_xxxvii(cc) and is_exceto_rpps:
                M_jur[mes]["XXXVII"] += vadeb - vacred; continue
            else:
                continue

            na2    = na[:2]
            na_mod = na[4:6]

            if func == "99":
                M_dep[mes]["XXIX"][col] += val
            elif is_exceto_rpps:
                if na2 in {"31","32","33"}:
                    M_dep[mes]["XVIII"][col] += val
                    if   na2 == "31": M_dep[mes]["pessoal"][col]              += val
                    elif na2 == "32": M_dep[mes]["XIX"][col]                  += val
                    elif na2 == "33": M_dep[mes]["outras_correntes_dep"][col] += val
                elif na2 in {"44","45","46"}:
                    M_dep[mes]["XXIII"][col] += val
                    if na2 == "44":
                        M_dep[mes]["investimentos"][col] += val
                    elif na2 == "45":
                        if   na_mod == "66": M_dep[mes]["XXIV"][col]     += val
                        elif na_mod == "64": M_dep[mes]["XXV"][col]      += val
                        elif na_mod == "63": M_dep[mes]["XXVI"][col]     += val
                        else:                M_dep[mes]["demais_inv"][col] += val
                    elif na2 == "46":
                        M_dep[mes]["XXVII"][col] += val
            else:
                if na2 in {"31","33"}:
                    M_dep[mes]["XXI"][col] += val
                elif na2 in {"44","45"} and not (na2 == "45" and na_mod in {"63","64","66"}):
                    M_dep[mes]["XXX"][col] += val

    max_mes = next((m for m in range(12, 0, -1)
                    if any(M_rec[m][k] != 0.0 for k in REC_KEYS)), 0)
    if max_mes == 0:
        max_mes = next((m for m in range(12, 0, -1)
                        if any(any(v != 0.0 for v in M_dep[m][k]) for k in DEP_KEYS)), 1)

    P_IV   = P_rec["I"]   - P_rec["II"]  - P_rec["III"]
    P_XIII = P_rec["VII"] - (P_rec["VIII"] + P_rec["IX"] + P_rec["X"] + P_rec["XI"] + P_rec["XII"])
    previsao = {k: round(P_rec[k], 2) for k in REC_KEYS}
    previsao.update({
        "IV":   round(P_IV, 2),
        "XIII": round(P_XIII, 2),
        "XVI":  round(P_IV + P_rec["V"] + P_XIII + P_rec["XIV"], 2),
        "XVII": round(P_IV + P_XIII, 2),
    })

    por_mes = {}
    for mes in range(1, max_mes + 1):
        por_mes[str(mes)] = {
            "rec": {k: round(M_rec[mes][k], 2) for k in REC_KEYS},
            "dep": {k: [round(M_dep[mes][k][i], 2) for i in range(3)] for k in DEP_KEYS},
            "jur": {"XXXVI": round(M_jur[mes]["XXXVI"], 2),
                    "XXXVII": round(M_jur[mes]["XXXVII"], 2)},
        }

    log.info(f"  Resultado Primário/Nominal: max_mes={max_mes}, meses={list(por_mes.keys())}")
    return {"max_mes": max_mes, "previsao": previsao, "por_mes": por_mes}


def init_oracle():
    import oracledb
    if CLIENT_PATH:
        log.info(f"Inicializando thick mode -> {CLIENT_PATH}")
        oracledb.init_oracle_client(lib_dir=CLIENT_PATH)
    else:
        log.info("Usando thin mode (sem Oracle Client local)")
    return oracledb


def run():
    try:
        oracledb = init_oracle()
    except ImportError:
        raise ImportError("Execute: pip install oracledb")

    if not DB_USER or not DB_PASSWORD:
        raise ValueError("DB_USER e DB_PASSWORD precisam estar definidos no .env")

    log.info(f"Conectando ao Oracle -> {DB_DSN}  [schema: {SCHEMA_ANO}]")

    pool = oracledb.create_pool(
        user=DB_USER, password=DB_PASSWORD, dsn=DB_DSN,
        min=DB_MIN, max=DB_MAX, increment=DB_INC,
    )

    with pool.acquire() as conn:
        log.info("Conexao estabelecida. Iniciando extracao...")
        with conn.cursor() as cur:
            for item in QUERIES:
                log.info(f"Extraindo -> {item['file']}")
                try:
                    data = fetch(cur, resolve_query(item))
                    if item.get("transform") == "rcl":
                        D_obj = build_rcl_data(data)
                        save_rcl_gz(D_obj)
                        upsert_rcl_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "restos_a_pagar":
                        registros = build_restos_a_pagar_data(data)
                        save_restos_a_pagar_gz(registros)
                        upsert_restos_a_pagar_supabase(registros)
                        save_json(item["file"], data)
                    elif item.get("transform") == "resultado_primario_nominal":
                        D_obj = build_resultado_primario_nominal_data(data)
                        save_resultado_primario_nominal_gz(D_obj)
                        upsert_resultado_primario_nominal_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item.get("transform") == "poupanca_corrente":
                        D_obj = build_poupanca_corrente_data(data)
                        save_poupanca_corrente_gz(D_obj)
                        upsert_poupanca_corrente_supabase(D_obj)
                        save_json(item["file"], data)
                    elif item["file"] == "receita.json":
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                        upsert_receita_supabase(data)
                    elif item["file"] == "despesa.json":
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                        upsert_despesa_supabase(data)
                    else:
                        save_json(item["file"], data)
                        save_json_gz(item["file"], data)
                except Exception as e:
                    log.error(f"  Erro em {item['file']}: {type(e).__name__}: {e}")
                    import traceback
                    traceback.print_exc()

    pool.close()
    log.info("ETL concluido com sucesso.")


if __name__ == "__main__":
    run()
